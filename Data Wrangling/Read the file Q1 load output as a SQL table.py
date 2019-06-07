# -*- coding: utf-8 -*-
"""
Created on Fri Jun  7 09:57:54 2019

@author: Training29
"""

# -*- coding: utf-8 -*-
"""
Created on Thu Jun  6 21:02:59 2019
@author: hzhou
"""

##### 1.
import pyodbc
import sqlalchemy
from datetime import datetime, time
from openpyxl import load_workbook
from math import radians, sin, cos, acos
values = []
conn = pyodbc.connect('Driver={SQL Server Native Client 11.0};'
                      'Server=SSINTR02;'
                      'Database=T5_Vijayakumar;'
                      'uid=T5_Vijayakumar;'
                      'pwd=vijay207$;')

engine=sqlalchemy.create_engine('mssql+pyodbc:///?odbc_connect={}'.format(conn))
cursor = conn.cursor()
cursor.execute('truncate TABLE dbo.Attribute')
i=0
for row in load_workbook('Q1.xlsx').worksheets[0].iter_rows():
    if i!=0:
        row = [cell.value if cell.value is not None else '' for cell in row] 
        timedelta = datetime.strptime( row[1], '%m/%d/%Y %H:%M')-datetime.strptime( row[0], '%m/%d/%Y %H:%M')
        duration=timedelta.days * 24 * 3600 + timedelta.seconds
        isWeekend = 1 if datetime.strptime( row[0], '%m/%d/%Y %H:%M').weekday()>=5 else 0
        hour_of_day=datetime.strptime( row[0], '%m/%d/%Y %H:%M').hour
        dist = 6371.01 * acos(sin(row[2])*sin(row[4]) + cos(row[2])*cos(row[4])*cos(row[3] - row[5]))
     
        cursor.execute("INSERT INTO dbo.Attribute(StartTime,EndTime,Start_Latitude,Start_Longitude,End_Latitude,End_Longitude,duration,is_Weekend,hour_of_the_day,distance) VALUES (?,?,?,?,?,?,?,?,?,?)" , row[0],row[1],row[2],row[3],row[4],row[5],duration,isWeekend,hour_of_day,dist)
    i=i+1 
conn.commit()
conn.close()